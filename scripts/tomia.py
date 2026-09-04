#!/usr/bin/env python3
"""nguoilonbinhan.meschool - boc file Tomia trong kho 'kho-tho' -> bang tomia_ngay.

Chay tren GitHub Actions. Khoa lay tu bien moi truong:
  SUPABASE_URL : https://xxx.supabase.co
  SUPABASE_KEY : service_role key

Cach doc sheet "01. TONG QUAN THEO TRUONG":
  - dong 3 (index 2): tieu de ngay "Ngay 01/06 (Mon)", moi ngay 1 khoi 16 cot, bat dau cot 2
  - dong 5 (index 4): ten 16 chi so trong moi khoi
  - dong 6 tro di   : moi dong = 1 truong; dong cuoi co ma_truong = 'TONG' -> bo
Nam lay tu ten tep (vd 01-06-2026_13-07-2026_xxx.xlsx).
File moi ghi de file cu qua khoa (ngay, ma_truong).
"""
import io, os, re, json, datetime, urllib.request, urllib.error

import openpyxl

SB_URL = os.environ["SUPABASE_URL"].rstrip("/")
SB_KEY = os.environ["SUPABASE_KEY"]
BUCKET = "kho-tho"
PREFIX = "tomia/"
SHEET = "01. TỔNG QUAN THEO TRƯỜNG"

CHI_SO = ["di_hoc", "vang", "den_muon", "ve_muon",
          "tre_dan_thuoc", "cu_thuoc", "da_uong",
          "gy_moi", "gy_dang_xl", "gy_hoan_tat", "gy_mo_lai",
          "vh_dang_xl", "vh_da_dong", "vh_cao", "vh_tb", "vh_thap"]


def goi(url, method="GET", body=None, headers=None, raw=False):
    h = {"apikey": SB_KEY, "Authorization": "Bearer " + SB_KEY}
    h.update(headers or {})
    data = None
    if body is not None:
        data = json.dumps(body).encode()
        h["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=h, method=method)
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read() if raw else r.read().decode()


def danh_sach_tep():
    out = goi(f"{SB_URL}/storage/v1/object/list/{BUCKET}", "POST",
              {"prefix": PREFIX, "limit": 200,
               "sortBy": {"column": "created_at", "order": "asc"}})
    # API tra ve ten tuong doi so voi prefix -> gan lai duong dan day du
    return [PREFIX + o["name"] for o in json.loads(out)
            if o.get("name", "").endswith(".xlsx")]


def nam_tu_ten(ten):
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", ten)
    return (int(m.group(3)), int(m.group(2))) if m else (datetime.date.today().year, 1)


def so(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def boc(ten_tep, noi_dung):
    wb = openpyxl.load_workbook(io.BytesIO(noi_dung), read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"  bo qua {ten_tep}: khong co sheet tong quan")
        return []
    ws = wb[SHEET]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if len(rows) < 6:
        return []
    nam, thang_dau = nam_tu_ten(ten_tep)

    cot_ngay = []
    for c, v in enumerate(rows[2]):
        if isinstance(v, str) and v.strip().startswith("Ngày"):
            m = re.search(r"(\d{2})/(\d{2})", v)
            if not m:
                continue
            ngay, thang = int(m.group(1)), int(m.group(2))
            n = nam + 1 if thang < thang_dau else nam   # bao ve truong hop qua nam
            try:
                cot_ngay.append((c, datetime.date(n, thang, ngay)))
            except ValueError:
                pass

    ket_qua = []
    for r in rows[5:]:
        ma = (r[0] or "").strip() if isinstance(r[0], str) else r[0]
        if not ma or str(ma).upper().startswith("TỔNG"):
            continue
        ten = r[1] if len(r) > 1 else None
        for c, d in cot_ngay:
            khoi = r[c:c + 16]
            gia_tri = {k: so(khoi[i]) if i < len(khoi) else None
                       for i, k in enumerate(CHI_SO)}
            if all(v is None for v in gia_tri.values()):
                continue
            gia_tri.update({"ngay": str(d), "ma_truong": str(ma),
                            "ten_truong": ten, "tep": ten_tep})
            ket_qua.append(gia_tri)
    return ket_qua


def ghi(dong):
    for i in range(0, len(dong), 500):
        goi(f"{SB_URL}/rest/v1/tomia_ngay?on_conflict=ngay,ma_truong", "POST",
            dong[i:i + 500],
            {"Prefer": "resolution=merge-duplicates,return=minimal"})


def main():
    teps = danh_sach_tep()
    print(f"Tim thay {len(teps)} tep Tomia trong kho")
    tong = 0
    for ten in teps:
        try:
            noi_dung = goi(f"{SB_URL}/storage/v1/object/{BUCKET}/{ten}", raw=True)
            dong = boc(ten, noi_dung)
            if dong:
                ghi(dong)
                tong += len(dong)
            print(f"  {ten}: {len(dong)} dong")
        except Exception as e:                    # mot tep loi khong lam sap ca lan chay
            print(f"  {ten}: LOI {type(e).__name__} {e}")
    print(f"Xong. Ghi {tong} dong vao tomia_ngay.")


if __name__ == "__main__":
    main()
